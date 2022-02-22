from sklearn.neural_network import MLPClassifier
from sklearn import svm
from sklearn import neighbors
from sklearn import linear_model


from openpyxl import load_workbook


# 统计预测准确率
def predict_accuracy(pred_y, real_y):
    accur = 0
    for i in range(len(pred_y)):
        if pred_y[i] == real_y[i]:
            accur += 1
    return accur/len(pred_y)


# 读取文件
def read_file(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    max_row = ws.max_row
    x = []
    for k in range(125):
        x.append([])
    y = []
    for i in range(2, max_row+1):
        y.append(ws['B{}'.format(i)].value if ws['B{}'.format(i)].value is not None else 0)
        x[0].append(ws['A{}'.format(i)].value if ws['A{}'.format(i)].value is not None else 0)
        for j in range(3, 127):
            x[j-2].append(ws.cell(row=i, column=j).value if ws.cell(row=i, column=j).value is not None else 0)

    return x, y


def main():
    train_x, train_y = read_file("happiness_train.xlsx")
    predict_x, predict_y = read_file("happiness_predict.xlsx")

    # 逻辑回归
    reg = linear_model.Lasso(alpha=0.1)
    reg.fit(train_x, train_y)
    accuracy = predict_accuracy(reg.predict(predict_x), predict_y)
    print(accuracy)

    # K近邻
    reg = neighbors.KNeighborsClassifier()
    reg.fit(train_x, train_y)
    accuracy = predict_accuracy(reg.predict(predict_x), predict_y)
    print(accuracy)

    # 支持向量机
    reg = svm.SVC()
    reg.fit(train_x, train_y)
    accuracy = predict_accuracy(reg.predict(predict_x), predict_y)
    print(accuracy)

    # 神经网络
    reg = MLPClassifier()
    reg.fit(train_x, train_y)
    accuracy = predict_accuracy(reg.predict(predict_x), predict_y)
    print(accuracy)


if __name__ == "__main__":
    main()

